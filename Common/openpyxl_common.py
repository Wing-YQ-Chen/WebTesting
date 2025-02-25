from openpyxl.worksheet.worksheet import *
from openpyxl.cell.cell import *
from openpyxl.styles import *
from Common.utils import *
import openpyxl
import os
import queue
import copy


def extract_input_data(input_file_path: str, header_row: int = 1, content_start_row: int = 2) -> queue.Queue:
    input_wk = openpyxl.load_workbook(input_file_path, True)
    input_ws = input_wk.worksheets[0]

    input_header = input_ws[f"A{header_row}":input_ws.cell(1, input_ws.max_column).column_letter + "1"]
    input_datas = input_ws[f"A{content_start_row}":input_ws.cell(1, input_ws.max_column).column_letter + input_ws.max_row.__str__()]

    input_table_q = queue.Queue()
    input_case = {}
    for r in range(input_datas.__len__()):
        for c in range(input_header[0].__len__()):
            input_case[input_header[0][c].value.__str__().strip()] = input_datas[r][c].value
        input_table_q.put(copy.copy(input_case))
        input_case = {}
    input_wk.close()
    return input_table_q


def openpyxl_merge_cell(sheet_object: Worksheet, start_cell: Cell, resize_y: int = 0, resize_x: int = 0):
    if resize_y < 0: resize_y = 0
    if resize_x < 0: resize_x = 0
    end_column = start_cell.column + resize_x
    end_row = start_cell.row + resize_y
    sheet_object.merge_cells(start_row=start_cell.row, start_column=start_cell.column, end_column=end_column,
                             end_row=end_row)


def openpyxl_find(cell_table, find_re: str, raise_e_msg: str = ""):
    for cell_list in cell_table:
        for cell in cell_list:
            if re.findall(find_re, cell.value.__str__().strip(), re.IGNORECASE):
                return cell
    if raise_e_msg:
        raise ValueError(raise_e_msg)
    else:
        return None


def create_report_xl(report_template_path, report_dir_path, xl_name, keep_sheet: list):
    report_wk = openpyxl.load_workbook(report_template_path)
    keep_name_list = []
    for sheet in keep_sheet:
        var_type = type(sheet)
        if var_type is int:
            keep_name_list.append(report_wk.worksheets[sheet].title)
        elif var_type is str:
            keep_name_list.append(sheet)

    for sheet in report_wk.worksheets:
        if sheet.title not in keep_name_list: report_wk.remove(sheet)

    if not os.path.exists(report_dir_path): os.makedirs(report_dir_path)
    report_path = os.path.join(report_dir_path, "{} - {}.xlsx".format(xl_name, get_timestamp_str()))
    report_wk.save(report_path)
    return report_path


def get_value_from_cell_table(cell_table):
    cell_value_table = list()
    cell_value_list = list()
    for cell_list in cell_table:
        cell_value_list = list()
        for cell in cell_list:
            cell_value_list.append(cell.value.__str__().strip() if cell.value else "")
        cell_value_table.append(list(cell_value_list))
    return cell_value_table if cell_value_table.__len__() > 1 else cell_value_list


def set_value_to_range(xl_sheet, start_cell_coordinate, value, highlight=None, red_keyword_list: list = [], green_keyword_list: list = []):
    """
    :param xl_workbook: sheet name or index
    :param sheet: int or str
    :param start_cell_coordinate: 'B2'
    :param value: str or list [
        [1,2,3],
        [4],
        [5,6,7,8,9]
    ]
    :param highlight:
    :param red_keyword_list: ['This word will be set red font']
    :return:

    橙色 #E87722
    黑蓝 #183028
    浅灰 #F2F2F2
    """

    if not value: return

    red_font = Font(name="DengXian", color="FF0000")
    green_font = Font(name="DengXian", color="00B050")
    orange_fill = PatternFill("solid", fgColor="E87722")
    yellow_fill = PatternFill("solid", fgColor="FFFF00")

    # xl_sheet = xl_workbook.worksheets[sheet]
    fill = PatternFill("solid", fgColor="E87722")
    side = Side(border_style='thin', color='00000000')
    border = Border(left=side, right=side, top=side, bottom=side)
    alignment = Alignment(horizontal='center', vertical='center')
    DEFAULT_style = NamedStyle(name="PYXL_DEFAULT", fill=fill, font=Font(color='00183028'), border=border, alignment=alignment)
    RED_style = NamedStyle(name="PYXL_RED", fill=fill, font=red_font, border=border, alignment=alignment)

    # start_cell = xl_sheet[start_cell_coordinate + (xl_sheet.max_row + 1).__str__()]
    start_cell = xl_sheet[start_cell_coordinate]
    first_row_flag = True
    first_col_flag = True
    row_number = start_cell.row
    col_number = start_cell.column
    if type(value) in [list, tuple]:
        value = list(value)
        if type(value[0]) not in [list, tuple]: value = [value]
        for row_index in range(value.__len__()):
            if first_row_flag:
                first_row_flag = False
                row_number = start_cell.row
            else:
                first_col_flag = True
                row_number += 1
                col_number = start_cell.column

            for col_index in range(value[row_index].__len__()):
                if first_col_flag:
                    first_col_flag = False
                    col_number = start_cell.column
                else:
                    col_number += 1
                current_cell = xl_sheet.cell(row_number, col_number)
                current_cell.value = value[row_index][col_index].__str__() if value[row_index][col_index] else ''
                if highlight and highlight[row_index:row_index + 1] and highlight[row_index][col_index:col_index + 1]:
                    temp_var = highlight[row_index][col_index]
                    if type(temp_var) is Font:
                        current_cell.font = temp_var
                    elif type(temp_var) is PatternFill:
                        current_cell.fill = temp_var
                # current_cell.style = DEFAULT_style.name if DEFAULT_style.name in xl_workbook.named_styles else DEFAULT_style
                if current_cell.value in red_keyword_list:
                    current_cell.font = red_font
                elif current_cell.value in green_keyword_list:
                    current_cell.font = green_font

    else:
        start_cell.value = value.__str__()
        if highlight: start_cell.font = highlight
        if start_cell.value in red_keyword_list:
            start_cell.font = red_font
        elif start_cell.value in green_keyword_list:
            start_cell.font = green_font


def cell_merge(ws, start_row, start_col, end_row, end_col):
    if start_row <= end_row and start_col <= end_col:
        ws.merge_cells(start_row=start_row, start_column=start_col, end_row=end_row, end_column=end_col)
